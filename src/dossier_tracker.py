"""
Dossier Tracker Académico v4
Rastreo diario de convocatorias con enriquecimiento automático SCImago / Latindex / CAICYT.
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
import smtplib
import ssl
import sys
import unicodedata
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urljoin, urlparse

import concurrent.futures
import requests
import yaml
from bs4 import BeautifulSoup
from dateparser import parse as parse_date
from dateparser.search import search_dates

# ─── Rutas ───────────────────────────────────────────────────────────────────

BASE_DIR      = Path(__file__).resolve().parents[1]
CONFIG_DIR    = BASE_DIR / "config"
RANKINGS_DIR  = CONFIG_DIR / "rankings"
# Buscar CSV de rankings en múltiples ubicaciones (config/rankings/ y raíz del repo)
RANKINGS_SEARCH_DIRS = [RANKINGS_DIR, BASE_DIR]
DATA_DIR      = BASE_DIR / "data"
REPORTS_DIR   = BASE_DIR / "reports"
DOCS_DATA_DIR = BASE_DIR / "docs" / "data"
URL_CACHE_PATH = DATA_DIR / "url_cache.json"     # ISSN → URL real de la revista

# ─── Entorno ──────────────────────────────────────────────────────────────────

DEFAULT_TIMEOUT      = int(os.getenv("HTTP_TIMEOUT_SECONDS", "15"))
MAX_WORKERS          = int(os.getenv("MAX_WORKERS", "8"))
MAX_LINKS_PER_SOURCE = int(os.getenv("MAX_LINKS_PER_SOURCE", "12"))
MIN_SCORE            = float(os.getenv("MIN_SCORE", "3.0"))
ONLY_OPEN            = os.getenv("ONLY_OPEN", "true").strip().lower() == "true"
CREATE_GITHUB_ISSUE  = os.getenv("CREATE_GITHUB_ISSUE", "false").strip().lower() == "true"
CURRENT_YEAR         = datetime.now(timezone.utc).year

USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36 dossier-latam-tracker/4.0"
)

# ─── Terminología ─────────────────────────────────────────────────────────────

CALL_TERMS = [
    "convocatoria", "call for papers", "call for paper", "llamado", "llamada",
    "dossier", "dosier", "chamada", "submissions", "submission",
    "próximos números", "proximos numeros", "announcement", "special issue",
    "monográfico", "monografico", "tema libre", "thematic issue", "open call",
    "recepción de artículos", "recepcion de articulos", "número regular",
    "numero regular", "se recibirán", "se recibiran",
]

PERMANENT_TERMS = [
    "convocatoria permanente", "recepción continua", "recepcion continua",
    "flujo continuo", "publicación continua", "publicacion continua",
    "continuous submission", "rolling submission", "open submissions",
    "tema libre", "envío permanente", "envio permanente", "always open",
]

DEADLINE_TERMS = [
    "hasta", "fecha límite", "fecha limite", "deadline", "cierre",
    "recepción de artículos", "recepcion de articulos", "envío", "envio",
    "submissions", "se recibirán", "se recibiran", "abierta la recepción",
    "abierta la recepcion", "abierto el llamado", "chamada", "submissão",
]

# ─── Reglas temáticas ─────────────────────────────────────────────────────────

THEME_RULES: dict[str, list[str]] = {
    "dirección escolar": [
        "dirección escolar", "direccion escolar", "gestión escolar",
        "gestion escolar", "school leadership", "educational leadership",
        "school management", "liderazgo escolar", "liderazgo educativo",
        "director escolar", "directivos escolares", "gobierno de la escuela",
    ],
    "educación": [
        "educación", "educacion", "pedagogía", "pedagogia", "enseñanza",
        "currículo", "curriculo", "escuela", "escolar", "trabajo docente",
        "didáctica", "didactica",
    ],
    "sociología": [
        "sociología", "sociologia", "sociedad", "juventudes", "adolescencias",
        "territorio", "desigualdad", "pobreza", "inclusión", "inclusion",
    ],
    "política": [
        "política", "politica", "políticas educativas", "politicas educativas",
        "democracia", "derechas", "gobierno", "estado", "poder",
    ],
    "humanidades": [
        "humanidades", "historia", "filosofía", "filosofia",
        "cultura", "lenguaje", "lectura", "escritura",
    ],
    "ciencias sociales": [
        "ciencias sociales", "social sciences", "movimientos sociales",
        "territorios", "américa latina", "america latina",
        "latinoamérica", "latinoamerica", "iberoamérica", "iberoamerica",
    ],
}

QUALITY_BONUS = {
    "Q1": 22, "Q2": 16, "Q3": 10, "Q4": 6,
    "NBRA": 8, "DOAJ": 6, "SciELO": 6, "RedALyC": 5, "Latindex": 3,
}


# ─── Modelos ──────────────────────────────────────────────────────────────────

@dataclass
class Item:
    source_name: str
    source_url: str
    title: str
    url: str
    region: str
    language: str
    tags: list[str]
    summary: str
    score: float
    deadline_text: str | None
    deadline_iso: str | None
    publication_date_text: str | None
    publication_date_iso: str | None
    status: str
    country: str
    city: str
    latitude: float | None
    longitude: float | None
    call_mode: str
    source_kind: str
    themes: list[str]
    strategic_score: int
    urgency: str
    days_left: int | None
    dossier_topic: str
    quartile: str | None
    quality_label: str | None
    quality_source: str | None
    sjr: float | None = None
    h_index: int | None = None
    open_access: bool = False
    issn: str | None = None
    latindex_catalogada: bool = False
    latindex_subtemas: list[str] = field(default_factory=list)
    call_type: str = "convocatoria"  # dossier | especial | convocatoria | continua | cerrada
    url_confidence: str = ""         # high | medium | low | ""

    @property
    def fingerprint(self) -> str:
        base = f"{normalize_space(self.title).lower()}|{self.url.strip().lower()}"
        return re.sub(r"\s+", " ", base).strip()


# ─── Utilidades de texto ──────────────────────────────────────────────────────

def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def normalize_key(text: str) -> str:
    text = normalize_space(text).lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text


def clean_issn(raw: str) -> list[str]:
    """Extrae todos los ISSN de un string, devuelve lista normalizada sin guión."""
    found = re.findall(r"\d{4}-?\d{3}[\dXx]", raw)
    return [x.replace("-", "").upper() for x in found]


# ─── Carga de configuración ──────────────────────────────────────────────────

def load_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def load_seen(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        with path.open("r", encoding="utf-8") as f:
            return {str(x) for x in json.load(f)}
    except Exception:
        return set()


def save_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def read_keywords() -> dict[str, list[str]]:
    data = load_yaml(CONFIG_DIR / "keywords.yml")
    kw = data.get("keywords", {})
    return {
        "strong":  [x.lower() for x in kw.get("strong", [])],
        "medium":  [x.lower() for x in kw.get("medium", [])],
        "exclude": [x.lower() for x in kw.get("exclude", [])],
    }


# ─── Rankings: SCImago ───────────────────────────────────────────────────────

def load_scimago_db(rankings_dir: Path) -> dict[str, dict[str, Any]]:
    """
    Carga todos los CSV de SCImago en rankings_dir.
    Retorna dict: issn_sin_guion_mayus → {quartile, sjr, h_index, open_access, title, country, rank}
    Si un ISSN aparece en varias categorías, prevalece el cuartil más alto (Q1 > Q2 > …).
    """
    QUARTILE_ORDER = {"Q1": 0, "Q2": 1, "Q3": 2, "Q4": 3}
    db: dict[str, dict[str, Any]] = {}

    csv_files = []
    for search_dir in RANKINGS_SEARCH_DIRS:
        csv_files.extend(search_dir.glob("scimagojr_*.csv"))   # guiones_bajos
        csv_files.extend(search_dir.glob("scimagojr *.csv"))   # con espacios
        csv_files.extend(search_dir.glob("scimagojr*.csv"))    # cualquier variante
    csv_files = list({f.resolve() for f in csv_files})  # deduplicar
    if not csv_files:
        dirs_str = ", ".join(str(d) for d in RANKINGS_SEARCH_DIRS)
        print(f"[INFO] No se encontraron CSV de SCImago en: {dirs_str}", file=sys.stderr)
        return db

    for csv_path in csv_files:
        try:
            text = csv_path.read_text(encoding="utf-8-sig")
            reader = csv.DictReader(io.StringIO(text), delimiter=";")
            for row in reader:
                issns_raw = row.get("Issn", "") or ""
                quartile  = (row.get("SJR Quartile") or "").strip()
                if quartile not in QUARTILE_ORDER:
                    continue
                title   = normalize_space(row.get("Title", ""))
                sjr_raw = (row.get("SJR") or "0").replace(",", ".")
                try:
                    sjr = float(sjr_raw)
                except ValueError:
                    sjr = 0.0
                h_raw = (row.get("H index") or "0").strip()
                try:
                    h_index = int(h_raw)
                except ValueError:
                    h_index = 0
                open_access = (row.get("Open Access") or "").strip().lower() == "yes"
                country  = normalize_space(row.get("Country", ""))
                rank_raw = (row.get("Rank") or "9999").strip()
                try:
                    rank = int(rank_raw)
                except ValueError:
                    rank = 9999

                record = {
                    "quartile": quartile,
                    "sjr": sjr,
                    "h_index": h_index,
                    "open_access": open_access,
                    "title": title,
                    "country": country,
                    "rank": rank,
                }
                for issn in clean_issn(issns_raw):
                    existing = db.get(issn)
                    if existing is None:
                        db[issn] = record
                    else:
                        if QUARTILE_ORDER[quartile] < QUARTILE_ORDER.get(existing["quartile"], 9):
                            db[issn] = record
        except Exception as exc:
            print(f"[WARN] Error leyendo {csv_path.name}: {exc}", file=sys.stderr)

    print(f"[INFO] SCImago DB: {len(db)} registros cargados desde {len(csv_files)} CSV.", file=sys.stderr)
    return db


# ─── Rankings: Latindex ───────────────────────────────────────────────────────

def load_latindex_db(rankings_dir: Path) -> dict[str, dict[str, Any]]:
    """
    Carga todos los CSV Indice_temas*.csv del catálogo Latindex.
    Retorna dos índices combinados en un dict:
      issn_sin_guion → {catalogada, country, subtemas}
    """
    db: dict[str, dict[str, Any]] = {}

    csv_files = []
    for search_dir in RANKINGS_SEARCH_DIRS:
        csv_files.extend(search_dir.glob("Indice_temas*.csv"))
        csv_files.extend(search_dir.glob("indice_temas*.csv"))
        csv_files.extend(search_dir.glob("Indice temas*.csv"))  # con espacios
        csv_files.extend(search_dir.glob("indice temas*.csv"))
    csv_files = list({f.resolve() for f in csv_files})  # deduplicar
    if not csv_files:
        print("[INFO] No se encontraron CSV de Latindex en config/rankings/.", file=sys.stderr)
        return db

    for csv_path in csv_files:
        try:
            text = csv_path.read_text(encoding="utf-8-sig")
            reader = csv.DictReader(io.StringIO(text), delimiter=";")
            for row in reader:
                issn_e   = normalize_space(row.get("issn_e", ""))
                issn_l   = normalize_space(row.get("issn_l", ""))
                issn_imp = normalize_space(row.get("issn_imp", ""))
                catalogada_raw = (row.get("catalogada") or "0").strip()
                catalogada = catalogada_raw == "1"
                country  = normalize_space(row.get("nombre_largo", ""))
                subtemas_raw = row.get("subtemas", "") or ""
                subtemas = [s.strip() for s in subtemas_raw.split(",") if s.strip()]

                record = {"catalogada": catalogada, "country": country, "subtemas": subtemas}
                for issn in [issn_e, issn_l, issn_imp]:
                    for clean in clean_issn(issn):
                        if clean:
                            existing = db.get(clean)
                            if existing is None or (catalogada and not existing["catalogada"]):
                                db[clean] = record
        except Exception as exc:
            print(f"[WARN] Error leyendo {csv_path.name}: {exc}", file=sys.stderr)

    print(f"[INFO] Latindex DB: {len(db)} registros cargados.", file=sys.stderr)
    return db


# ─── Enriquecimiento de calidad ──────────────────────────────────────────────

def enrich_source_quality(
    source: dict[str, Any],
    scimago_db: dict[str, dict[str, Any]],
    latindex_db: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    """
    Retorna un dict con campos de calidad enriquecidos para una fuente.
    Prioridad: SCImago > CAICYT-CONICET (manual) > Latindex > sin dato.
    """
    issn_raw = source.get("issn", "")
    issns = clean_issn(issn_raw) if issn_raw else []

    # 1. Buscar en SCImago
    scimago_record: dict[str, Any] | None = None
    for issn in issns:
        rec = scimago_db.get(issn)
        if rec:
            scimago_record = rec
            break

    # 2. Buscar en Latindex
    latindex_record: dict[str, Any] | None = None
    for issn in issns:
        rec = latindex_db.get(issn)
        if rec:
            latindex_record = rec
            break

    result: dict[str, Any] = {
        "quartile":            source.get("quartile"),
        "quality_label":       source.get("quality_label"),
        "quality_source":      source.get("quality_source"),
        "sjr":                 None,
        "h_index":             None,
        "open_access":         False,
        "latindex_catalogada": False,
        "latindex_subtemas":   [],
    }

    if scimago_record:
        result["quartile"]     = scimago_record["quartile"]
        result["sjr"]          = scimago_record["sjr"]
        result["h_index"]      = scimago_record["h_index"]
        result["open_access"]  = scimago_record["open_access"]
        if not result["quality_source"]:
            result["quality_source"] = "SCImago 2024"
        if not result["quality_label"] or result["quality_label"] in ("Sin dato visible",):
            result["quality_label"] = scimago_record["quartile"]

    if latindex_record:
        result["latindex_catalogada"] = latindex_record["catalogada"]
        result["latindex_subtemas"]   = latindex_record["subtemas"]
        if not result["quality_label"] or result["quality_label"] in ("Sin dato visible",):
            result["quality_label"] = "Latindex"
        if not result["quality_source"]:
            result["quality_source"] = "Latindex Catálogo"

    return result


# ─── HTTP ────────────────────────────────────────────────────────────────────

def get_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    return session


def fetch_html(session: requests.Session, url: str) -> str | None:
    try:
        response = session.get(url, timeout=DEFAULT_TIMEOUT)
        response.raise_for_status()
        ct = response.headers.get("content-type", "")
        if "text/html" not in ct and "application/xhtml+xml" not in ct:
            return None
        response.encoding = response.encoding or "utf-8"
        return response.text
    except Exception as exc:
        print(f"[WARN] No se pudo abrir {url}: {exc}", file=sys.stderr)
        return None



# ─── Resolución masiva de URLs reales — pipeline por capas ──────────────────
#
# Arquitectura:
#   Capa 1: OpenAlex /sources/issn:{issn}   → homepage_url  (fuente canónica por ISSN-L)
#   Capa 2: Crossref /journals/{issn}        → URL / homepage-URL
#   Capa 3: DOAJ /api/v2/search/journals/{issn} → bibjson.ref.homepage
#   Capa 4: SCImago page (fallback live)     → primer link externo
#
# Luego de obtener una URL candidata se valida que:
#   - El dominio responda (HTTP 200 o 30x)
#   - La página mencione el título de la revista o su ISSN
#
# Campo `confidence`:
#   "high"   → OpenAlex/Crossref/DOAJ + validación superada
#   "medium" → fuente confiable pero validación no disponible o parcial
#   "low"    → fallback SCImago / sin validar
#   ""       → sin URL resuelta

_UA_BOT  = "dossier-tracker/4.0 (mailto:aguirre.elias.gonzalo@gmail.com)"
_UA_LITE = "dossier-tracker/4.0"


# ── Cache I/O ─────────────────────────────────────────────────────────────────

def load_url_cache(path: Path) -> dict[str, dict[str, str]]:
    """
    Caché: issn_key → {"url": str, "confidence": str, "source": str}
    Retorna dict vacío si no existe o está corrupto.
    """
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
        # Migrar formato viejo (str) → nuevo (dict)
        out: dict[str, dict[str, str]] = {}
        for k, v in raw.items():
            if isinstance(v, str):
                out[k] = {"url": v, "confidence": "medium", "source": "legacy"}
            elif isinstance(v, dict):
                out[k] = v
        return out
    except Exception:
        return {}


def save_url_cache(path: Path, cache: dict[str, dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


# ── Capa 1: OpenAlex Sources ──────────────────────────────────────────────────

def _openalex_source(session: requests.Session, issn: str) -> tuple[str, str] | None:
    """
    Consulta OpenAlex /sources/issn:{issn}.
    Retorna (homepage_url, tipo_fuente) o None.
    """
    for issn_val in [issn, issn.replace("-", "")]:
        try:
            r = session.get(
                f"https://api.openalex.org/sources/issn:{issn_val}",
                timeout=20, headers={"User-Agent": _UA_BOT},
            )
            if r.status_code == 200:
                d = r.json()
                url  = d.get("homepage_url") or ""
                kind = d.get("type") or ""
                if url and url.startswith("http"):
                    return url.rstrip("/"), kind
            elif r.status_code == 404:
                break
            elif r.status_code == 429:
                print(f"[WARN] OpenAlex rate limit — esperando 5s", file=sys.stderr)
                import time as _t; _t.sleep(5)
        except Exception as e:
            print(f"[WARN] OpenAlex {issn}: {e}", file=sys.stderr)
    return None


# ── Capa 2: Crossref ──────────────────────────────────────────────────────────

def _crossref_journal(session: requests.Session, issn: str) -> str | None:
    try:
        r = session.get(
            f"https://api.crossref.org/journals/{issn.strip()}",
            timeout=20, headers={"User-Agent": _UA_BOT},
        )
        if r.status_code == 429:
            print(f"[WARN] Crossref rate limit — esperando 5s", file=sys.stderr)
            import time as _t; _t.sleep(5)
            return None
        if r.status_code != 200:
            return None
        d = r.json().get("message", {})
        url = d.get("URL") or d.get("homepage-URL") or ""
        return url.rstrip("/") if url.startswith("http") else None
    except Exception as e:
        print(f"[WARN] Crossref {issn}: {e}", file=sys.stderr)
        return None


# ── Capa 3: DOAJ ──────────────────────────────────────────────────────────────

def _doaj_journal(session: requests.Session, issn: str) -> str | None:
    try:
        r = session.get(
            f"https://doaj.org/api/v2/search/journals/issn:{issn.strip()}",
            timeout=20, headers={"User-Agent": _UA_BOT},
        )
        if r.status_code != 200:
            return None
        results = r.json().get("results", [])
        if not results:
            return None
        ref = results[0].get("bibjson", {}).get("ref", {})
        url = ref.get("homepage") or ref.get("journal") or ""
        return url.rstrip("/") if url.startswith("http") else None
    except Exception as e:
        print(f"[WARN] DOAJ {issn}: {e}", file=sys.stderr)
        return None


# ── Validación de URL candidata ───────────────────────────────────────────────

def _validate_url(
    session: requests.Session,
    url: str,
    journal_title: str,
    issn: str,
) -> bool:
    """
    Verifica que la URL responda y mencione el título de la revista o su ISSN.
    Tolerante: si el servidor responde (incluso 40x) se acepta como válido parcialmente.
    """
    try:
        r = session.get(url, timeout=10, allow_redirects=True,
                        headers={"User-Agent": _UA_BOT})
        if r.status_code >= 500:
            return False
        text = r.text[:8000].lower()
        # Buscar título (primeras 3 palabras significativas) o ISSN
        title_words = [w for w in normalize_key(journal_title).split() if len(w) > 3][:3]
        issn_clean  = issn.replace("-", "")
        title_hit   = sum(1 for w in title_words if w in text) >= min(2, len(title_words))
        issn_hit    = issn_clean in text or issn in text
        return title_hit or issn_hit
    except Exception:
        return False


# ── Resolución principal ──────────────────────────────────────────────────────

def _resolve_one(
    session: requests.Session,
    issn: str,
    title: str,
    source_url: str,
) -> dict[str, str]:
    """
    Intenta resolver la URL real de una revista por ISSN usando el pipeline de capas.
    Retorna {"url": ..., "confidence": ..., "source": ...}
    """
    # Capa 1: OpenAlex Sources
    oa = _openalex_source(session, issn)
    if oa:
        url, kind = oa
        confidence = "high" if kind == "journal" else "medium"
        validated  = _validate_url(session, url, title, issn)
        if validated:
            return {"url": url, "confidence": confidence, "source": "openalex"}
        elif kind == "journal":
            # Confiamos en OpenAlex + journal aunque no valide
            return {"url": url, "confidence": "medium", "source": "openalex_unvalidated"}

    # Capa 2: Crossref
    cr_url = _crossref_journal(session, issn)
    if cr_url:
        validated = _validate_url(session, cr_url, title, issn)
        confidence = "high" if validated else "medium"
        return {"url": cr_url, "confidence": confidence, "source": "crossref"}

    # Capa 3: DOAJ
    dj_url = _doaj_journal(session, issn)
    if dj_url:
        validated = _validate_url(session, dj_url, title, issn)
        confidence = "high" if validated else "medium"
        return {"url": dj_url, "confidence": confidence, "source": "doaj"}

    # Sin URL resuelta por ninguna capa → registrar como vacío
    return {"url": "", "confidence": "", "source": ""}


def resolve_urls_batch(
    sources: list[dict[str, Any]],
    cache: dict[str, dict[str, str]],
    session: requests.Session,
    max_new: int = 400,
) -> dict[str, dict[str, str]]:
    """
    Resuelve URLs reales para fuentes SCImago/Latindex/ISSN-portal usando el
    pipeline por capas (OpenAlex → Crossref → DOAJ → scraping directo).
    Máximo `max_new` consultas nuevas por corrida (rate limit).
    Primera corrida: ~400 ISSNs ≈ 2-3 min. Siguientes: instantáneo desde caché.
    """
    import time as _time
    NEEDS = ("scimagojr.com", "latindex.org", "portal.issn.org")

    pending = [
        s for s in sources
        if s.get("issn")
        and any(x in s.get("url", "") for x in NEEDS)
        and s.get("call_mode_hint") != "catalog_only"
        and s["issn"].replace("-", "").strip() not in cache
    ][:max_new]

    if not pending:
        high = sum(1 for v in cache.values() if isinstance(v,dict) and v.get("confidence")=="high")
        print(f"[INFO] URL cache: {len(cache)} ISSNs — {high} alta confianza. Sin nuevas consultas.")
        return cache

    print(f"[INFO] Resolviendo {len(pending)} URLs (OpenAlex→Crossref→DOAJ→scraping)...")
    stats: dict[str, int] = {"high": 0, "medium": 0, "low": 0, "none": 0}

    for i, source in enumerate(pending):
        issn_key = source["issn"].replace("-", "").strip()
        result   = _resolve_one(session, source["issn"], source["name"], source["url"])
        cache[issn_key] = result
        stats[result["confidence"] or "none"] += 1

        if (i + 1) % 50 == 0:
            print(f"[INFO]   {i+1}/{len(pending)} — "
                  f"alta:{stats['high']} media:{stats['medium']} "
                  f"baja:{stats['low']} sin URL:{stats['none']}")
        # Mostrar primeros 3 resultados para diagnóstico
        if i < 3:
            print(f"[DEBUG] ISSN {source['issn']} → {result}", file=sys.stderr)
        _time.sleep(0.15)  # ~6-7 req/s, dentro de límites de las 3 APIs

    total_resolved = stats["high"] + stats["medium"] + stats["low"]
    print(f"[INFO] Resueltas: {total_resolved}/{len(pending)} "
          f"(alta:{stats['high']} media:{stats['medium']} baja:{stats['low']})")
    return cache


# ─── Resolución de URL real (Latindex / ISSN portal) ─────────────────────────

DOSSIER_TERMS = [
    "dossier", "dosier", "número especial", "numero especial",
    "special issue", "thematic issue", "monográfico", "monografico",
    "número temático", "numero tematico",
]
SPECIAL_TERMS = [
    "número especial", "numero especial", "special issue", "thematic issue",
    "monográfico", "monografico", "número temático",
]


_SKIP_DOMAINS = (
    "twitter.com","facebook.com","instagram.com","linkedin.com","youtube.com",
    "google.com","issn.org","latindex.org","doi.org","crossref.org","orcid.org",
    "wikipedia.org","academia.edu","researchgate.net",
)

def resolve_real_url(session: requests.Session, url: str, sid: str = "") -> str:
    """
    Intenta extraer la URL real de una revista desde fichas de Latindex o portal ISSN.
    SCImago retorna 403 — no se intenta nada con esas URLs.
    """
    is_latindex = "latindex.org/latindex/ficha" in url
    is_issn     = "portal.issn.org/resource/ISSN" in url
    if not (is_latindex or is_issn):
        return url  # SCImago u otras URLs → devolver tal cual sin intentar
    html = fetch_html(session, url)
    if not html:
        return url
    soup = BeautifulSoup(html, "html.parser")
    src_netloc = urlparse(url).netloc
    for a in soup.find_all("a", href=True):
        href = a.get("href","").strip()
        if not href.startswith("http"):
            continue
        if urlparse(href).netloc == src_netloc:
            continue
        if any(s in href for s in _SKIP_DOMAINS):
            continue
        return href
    return url


def infer_call_type(title: str, text: str, call_mode: str, status: str) -> str:
    """Clasifica el tipo de convocatoria detectada."""
    if call_mode == "permanent":
        return "continua"
    if status == "vencida":
        return "cerrada"
    blob = (title + " " + text).lower()
    if any(t in blob for t in DOSSIER_TERMS):
        return "dossier"
    if any(t in blob for t in SPECIAL_TERMS):
        return "especial"
    return "convocatoria"


def soup_text(soup: BeautifulSoup, max_chars: int = 5000) -> str:
    for tag in soup(["script", "style", "noscript", "svg", "img"]):
        tag.decompose()
    return normalize_space(soup.get_text(" ", strip=True))[:max_chars]


def find_candidate_links(source_url: str, soup: BeautifulSoup) -> list[tuple[str, str]]:
    parsed_source = urlparse(source_url)
    candidates: list[tuple[str, str]] = []
    seen: set[str] = set()

    page_title = normalize_space(soup.title.get_text(" ", strip=True) if soup.title else "")
    page_text  = soup_text(soup, max_chars=2500)
    page_blob  = f"{page_title} {page_text}".lower()
    if any(t in page_blob for t in CALL_TERMS + PERMANENT_TERMS):
        seen.add(source_url)
        candidates.append((page_title or source_url, source_url))

    for a in soup.find_all("a", href=True):
        href = normalize_space(a.get("href", ""))
        text = normalize_space(a.get_text(" ", strip=True))
        if not href:
            continue
        full_url = urljoin(source_url, href)
        if urlparse(full_url).netloc != parsed_source.netloc:
            continue
        target = f"{text} {full_url}".lower()
        if not any(t in target for t in CALL_TERMS + PERMANENT_TERMS):
            continue
        if full_url in seen:
            continue
        seen.add(full_url)
        candidates.append((text or full_url, full_url))
        if len(candidates) >= MAX_LINKS_PER_SOURCE:
            break

    return candidates


# ─── Extracción de fechas ─────────────────────────────────────────────────────

DATE_CONTEXT_PATTERNS = [
    re.compile(
        r"(?i)(?:hasta|fecha l[ií]mite|deadline|cierre(?: de inscripci[oó]n)?|"
        r"recepci[oó]n de art[ií]culos(?:.*?hasta)?|"
        r"env[ií]o(?:s)?(?: de propuestas)?(?:.*?hasta)?|"
        r"submissions? (?:until|by)|submiss[oõ]es? at[eé])\s*[:\-–]?\s*([^\n\.;]{4,90})"
    ),
    re.compile(r"(?i)se recibir[aá]n[^\n]{0,80}? al\s+([^\n\.;]{4,90})"),
    re.compile(r"(?i)abierta la recepci[oó]n[^\n]{0,80}? hasta\s+([^\n\.;]{4,90})"),
]


def parse_multilingual_date(text: str | None) -> datetime | None:
    if not text:
        return None
    dt = parse_date(text, languages=["es", "pt", "en"], settings={
        "PREFER_DAY_OF_MONTH": "first",
        "PREFER_DATES_FROM": "future",
        "RELATIVE_BASE": datetime.now(),
    })
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _context_windows(snippet: str) -> list[str]:
    low = snippet.lower()
    windows = []
    for term in DEADLINE_TERMS:
        start = 0
        while True:
            idx = low.find(term, start)
            if idx == -1:
                break
            windows.append(snippet[max(0, idx - 20): min(len(snippet), idx + 140)])
            start = idx + len(term)
    return windows


def extract_deadline(text: str) -> tuple[str | None, str | None]:
    snippet = normalize_space(text)
    for pattern in DATE_CONTEXT_PATTERNS:
        for match in pattern.finditer(snippet):
            candidate = normalize_space(match.group(1))
            candidate = re.split(r"(?i)(?:publicaci[oó]n|evaluaci[oó]n|proceso)", candidate)[0].strip()
            dt = parse_multilingual_date(candidate)
            if dt and dt.date().isoformat() >= datetime.now(timezone.utc).date().isoformat():
                return candidate, dt.date().isoformat()

    for window in _context_windows(snippet[:3500]):
        found = search_dates(window, languages=["es", "pt", "en"], settings={
            "PREFER_DATES_FROM": "future",
            "PREFER_DAY_OF_MONTH": "first",
            "RELATIVE_BASE": datetime.now(),
        })
        if found:
            today = datetime.now(timezone.utc)
            future = []
            for raw, dt in found:
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                dt = dt.astimezone(timezone.utc)
                if dt.date() >= today.date():
                    future.append((normalize_space(raw), dt))
            if future:
                raw, dt = min(future, key=lambda x: x[1])
                return raw, dt.date().isoformat()
    return None, None


def extract_publication_date(soup: BeautifulSoup, text: str) -> tuple[str | None, str | None]:
    for selector in ["time", "meta[property='article:published_time']",
                     "meta[name='citation_publication_date']", "meta[name='DC.Date.created']"]:
        node = soup.select_one(selector)
        if not node:
            continue
        raw = normalize_space(
            node.get("datetime") or node.get("content") or node.get_text(" ", strip=True)
        )
        if not raw:
            continue
        dt = parse_multilingual_date(raw)
        if dt:
            return raw, dt.date().isoformat()
    return None, None


# ─── Puntuación ──────────────────────────────────────────────────────────────

def keyword_score(title: str, text: str, tags: Iterable[str], keywords: dict[str, list[str]]) -> float:
    blob = f"{title} {text} {' '.join(tags)}".lower()
    strong_hits   = sum(1 for kw in keywords["strong"] if kw in blob)
    medium_hits   = sum(1 for kw in keywords["medium"] if kw in blob)
    exclude_hits  = sum(1 for kw in keywords["exclude"] if kw in blob)
    call_hits     = sum(1 for t  in CALL_TERMS       if t  in blob)
    permanent_hits= sum(1 for t  in PERMANENT_TERMS  if t  in blob)
    score = (
        min(strong_hits, 6)    * 2.8
        + min(medium_hits, 8)  * 0.8
        + min(call_hits, 4)    * 1.2
        + min(permanent_hits, 2) * 0.8
        - min(exclude_hits, 3) * 2.0
    )
    return round(score, 2)


def determine_call_mode(source: dict[str, Any], title: str, page_text: str, deadline_iso: str | None) -> str:
    hint = normalize_space(str(source.get("call_mode_hint", ""))).lower()
    blob = f"{title} {page_text}".lower()
    if hint == "permanent":
        return "permanent"
    if any(t in blob for t in PERMANENT_TERMS):
        return "permanent"
    if deadline_iso:
        return "deadline"
    return "unknown"


def determine_status(deadline_iso: str | None, call_mode: str) -> str:
    if call_mode == "permanent":
        return "permanente"
    if not deadline_iso:
        return "sin_fecha"
    today = datetime.now(timezone.utc).date().isoformat()
    return "abierta" if deadline_iso >= today else "vencida"


def compute_days_left(deadline_iso: str | None, call_mode: str) -> int | None:
    if call_mode == "permanent" or not deadline_iso:
        return None
    today  = datetime.now(timezone.utc).date()
    target = datetime.fromisoformat(deadline_iso).date()
    return (target - today).days


def compute_urgency(days_left: int | None, call_mode: str) -> str:
    if call_mode == "permanent":
        return "continua"
    if days_left is None:
        return "sin_fecha"
    if days_left < 0:
        return "vencida"
    if days_left <= 7:
        return "crítica"
    if days_left <= 21:
        return "próxima"
    return "abierta"


def infer_dossier_topic(title: str) -> str:
    cleaned = normalize_space(title)
    for pattern in [
        r"(?i)^convocatoria\s*[:\-–]?\s*", r"(?i)^call for papers\s*[:\-–]?\s*",
        r"(?i)^llamado\s*[:\-–]?\s*",       r"(?i)^announcement\s*[:\-–]?\s*",
        r"(?i)^dossier\s*[:\-–]?\s*",       r"(?i)^dosier\s*[:\-–]?\s*",
    ]:
        cleaned = re.sub(pattern, "", cleaned).strip()
    cleaned = re.sub(
        r"(?i)(convocatoria|call for papers|announcement|recepci[oó]n continua|flujo continuo)",
        "", cleaned,
    ).strip(" -–—:;,.")
    return cleaned or normalize_space(title)


def infer_themes(title: str, text: str, tags: Iterable[str]) -> list[str]:
    blob = f"{title} {text} {' '.join(tags)}".lower()
    themes = [t for t, terms in THEME_RULES.items() if any(term in blob for term in terms)]
    return themes or ["otras"]


def quality_bonus(quartile: str | None, quality_label: str | None) -> int:
    if quartile and quartile in QUALITY_BONUS:
        return QUALITY_BONUS[quartile]
    if quality_label and quality_label in QUALITY_BONUS:
        return QUALITY_BONUS[quality_label]
    return 0


def thematic_priority(themes: list[str]) -> int:
    weights = {
        "dirección escolar": 16, "política": 12, "educación": 10,
        "sociología": 10, "ciencias sociales": 10, "humanidades": 6, "otras": 2,
    }
    return sum(weights.get(t, 0) for t in themes[:3])


def strategic_score(item: Item) -> int:
    base         = min(int(round(item.score * 6.5)), 42)
    theme_bonus  = min(thematic_priority(item.themes), 24)
    q_bonus      = quality_bonus(item.quartile, item.quality_label)
    urgency_bonus = 0
    if item.call_mode == "deadline" and item.days_left is not None:
        if 0 <= item.days_left <= 14:   urgency_bonus = 8
        elif 15 <= item.days_left <= 45: urgency_bonus = 5
        else:                            urgency_bonus = 2
    elif item.call_mode == "permanent":
        urgency_bonus = 4
    return int(round(min(base + theme_bonus + q_bonus + urgency_bonus, 100)))


def summarize_text(text: str, max_len: int = 420) -> str:
    text = normalize_space(text)
    if len(text) <= max_len:
        return text
    cut = text[: max_len - 1]
    if " " in cut:
        cut = cut.rsplit(" ", 1)[0]
    return cut + "…"


def stale_year_guard(title: str, page_text: str, pub_iso: str | None, call_mode: str) -> bool:
    if call_mode == "permanent":
        return False
    years = [int(y) for y in re.findall(r"(20\d{2})", f"{title} {page_text[:500]}")]
    if years and max(years) < CURRENT_YEAR:
        return True
    if pub_iso:
        try:
            if datetime.fromisoformat(pub_iso).year < CURRENT_YEAR - 1:
                return True
        except Exception:
            pass
    return False


# ─── Parsing de ítems ─────────────────────────────────────────────────────────

def parse_item_from_page(
    source: dict[str, Any],
    url: str,
    html: str,
    keywords: dict[str, list[str]],
    quality_fields: dict[str, Any],
) -> Item | None:
    soup      = BeautifulSoup(html, "html.parser")
    title_node = (soup.find("h1") or soup.find("title") or soup.find(["h2", "h3"]))
    title      = normalize_space(title_node.get_text(" ", strip=True) if title_node else "") or url
    page_text  = soup_text(soup, max_chars=6000)

    score         = keyword_score(title, page_text, source.get("tags", []), keywords)
    deadline_text, deadline_iso = extract_deadline(page_text)
    pub_text, pub_iso           = extract_publication_date(soup, page_text)
    call_mode     = determine_call_mode(source, title, page_text, deadline_iso)
    status        = determine_status(deadline_iso, call_mode)

    if score < MIN_SCORE:
        return None
    if ONLY_OPEN and status in {"vencida", "sin_fecha"} and call_mode != "permanent":
        return None
    if stale_year_guard(title, page_text, pub_iso, call_mode):
        return None

    themes    = infer_themes(title, page_text, source.get("tags", []))
    days_left = compute_days_left(deadline_iso, call_mode)
    urgency   = compute_urgency(days_left, call_mode)
    dossier_topic = infer_dossier_topic(title)

    call_type = infer_call_type(title, page_text, call_mode, status)
    item = Item(
        source_name=source["name"], source_url=source["url"], title=title, url=url,
        region=source.get("region", ""), language=source.get("language", ""),
        tags=list(source.get("tags", [])), summary=summarize_text(page_text),
        score=score, deadline_text=deadline_text, deadline_iso=deadline_iso,
        publication_date_text=pub_text, publication_date_iso=pub_iso,
        status=status, country=source.get("country", ""), city=source.get("city", ""),
        latitude=source.get("latitude"), longitude=source.get("longitude"),
        call_mode=call_mode, source_kind=source.get("source_kind", "revista"),
        themes=themes, strategic_score=0, urgency=urgency, days_left=days_left,
        dossier_topic=dossier_topic,
        quartile=quality_fields.get("quartile"),
        quality_label=quality_fields.get("quality_label"),
        quality_source=quality_fields.get("quality_source"),
        sjr=quality_fields.get("sjr"),
        h_index=quality_fields.get("h_index"),
        open_access=quality_fields.get("open_access", False),
        issn=source.get("issn"),
        latindex_catalogada=quality_fields.get("latindex_catalogada", False),
        latindex_subtemas=quality_fields.get("latindex_subtemas", []),
        call_type=call_type,
        url_confidence=quality_fields.get('_url_confidence',''),
    )
    item.strategic_score = strategic_score(item)
    return item


# ─── Deduplicación ───────────────────────────────────────────────────────────

def quality_rank(item: Item) -> int:
    q = {"Q1": 14, "Q2": 13, "Q3": 12, "Q4": 11}
    if item.quartile and item.quartile in q:
        return q[item.quartile]
    l = {"NBRA": 8, "SciELO": 7, "RedALyC": 6, "DOAJ": 5, "Latindex": 4}
    return l.get(item.quality_label or "", 0)


def canonical_item_key(item: Item) -> str:
    if item.call_mode == "permanent":
        return "|".join([
            normalize_key(item.source_name),
            normalize_key(item.country),
            normalize_key(item.city),
            "permanent",
        ])
    return item.fingerprint


def item_rank(item: Item) -> tuple[int, int, int, int]:
    generic = {"avisos", "announcement", "announcements", "submissions", "about", "inicio"}
    non_generic = 1 if normalize_key(item.title) not in generic else 0
    return (item.strategic_score or 0, quality_rank(item), non_generic, -len(item.url or ""))


def collect_items(
    session: requests.Session,
    source: dict[str, Any],
    keywords: dict[str, list[str]],
    quality_fields: dict[str, Any],
    url_cache: dict[str, str] | None = None,
) -> list[Item]:
    items: list[Item] = []
    # Las fuentes catalog_only solo aparecen en el catálogo visual; no se scrapean.
    if source.get("call_mode_hint") == "catalog_only":
        return items
    # Resolver URL real: primero buscar en caché (Crossref/OpenAlex), luego scraping live.
    issn_key = (source.get("issn") or "").replace("-", "").strip()
    NEEDS = ("scimagojr.com", "latindex.org", "portal.issn.org")
    needs_resolve = any(x in source.get("url", "") for x in NEEDS)
    if needs_resolve:
        if url_cache and issn_key and issn_key in url_cache:
            cached_url = url_cache[issn_key]
            if not cached_url:
                return items   # ISSN procesado pero sin URL → saltear
            source = {**source, "url": cached_url}
        else:
            # Fallback: intentar extraer URL desde la página de metadatos en tiempo real
            real_url = resolve_real_url(session, source["url"])
            if real_url == source["url"]:
                return items   # No se pudo resolver → saltear
            source = {**source, "url": real_url}
    source_html = fetch_html(session, source["url"])
    if not source_html:
        return items

    source_soup = BeautifulSoup(source_html, "html.parser")
    candidates  = find_candidate_links(source["url"], source_soup) or [(source["name"], source["url"])]
    seen_urls: set[str] = set()

    for _, candidate_url in candidates:
        if candidate_url in seen_urls:
            continue
        seen_urls.add(candidate_url)
        html = source_html if candidate_url == source["url"] else fetch_html(session, candidate_url)
        if not html:
            continue
        qf_c = {**quality_fields, '_url_confidence': source.get('_url_confidence','')}
        item = parse_item_from_page(source, candidate_url, html, keywords, qf_c)
        if item is not None:
            items.append(item)

    deduped: dict[str, Item] = {}
    for item in items:
        key = canonical_item_key(item)
        current = deduped.get(key)
        if current is None or item_rank(item) > item_rank(current):
            deduped[key] = item
    return list(deduped.values())


def sort_items(items: list[Item]) -> list[Item]:
    def sort_key(item: Item) -> tuple[int, str, int, int, float]:
        status_rank = 0 if item.status == "abierta" else 1 if item.status == "permanente" else 3
        deadline    = item.deadline_iso or "9999-99-99"
        days        = item.days_left if item.days_left is not None else 999999
        q_rank      = {"Q1": 0, "Q2": 1, "Q3": 2, "Q4": 3}.get(item.quartile or "", 4)
        return (status_rank, deadline, days, q_rank, -item.score)
    return sorted(items, key=sort_key)


# ─── Reportes Markdown ───────────────────────────────────────────────────────

def render_item(item: Item) -> list[str]:
    lines = [
        f"### {item.title}", "",
        f"- Fuente: **{item.source_name}**",
        f"- País / ciudad: **{item.country} / {item.city}**",
        f"- Temas: **{', '.join(item.themes)}**",
        f"- Puntaje estratégico: **{item.strategic_score}**",
        f"- Estado: **{item.status}**",
    ]
    if item.quartile:
        lines.append(f"- Cuartil SCImago: **{item.quartile}** (SJR {item.sjr})")
    if item.quality_label:
        lines.append(f"- Calidad: **{item.quality_label}** · {item.quality_source or ''}")
    if item.latindex_catalogada:
        lines.append(f"- Latindex: catalogada ✔")
    if item.deadline_text:
        label = item.deadline_text + (f" ({item.deadline_iso})" if item.deadline_iso else "")
        lines.append(f"- Cierre: **{label}**")
    lines.append(f"- Enlace: {item.url}")
    lines.extend(["", item.summary, ""])
    return lines


def render_markdown(new_items: list[Item], all_items: list[Item]) -> str:
    today = datetime.now(timezone.utc).date().isoformat()
    lines = [
        f"# Dossier diario de convocatorias | {today}", "",
        f"- Nuevas: **{len(new_items)}** | Total relevantes: **{len(all_items)}**", "",
        "## Novedades del día", "",
    ]
    if new_items:
        for item in new_items:
            lines.extend(render_item(item))
    else:
        lines.append("No aparecieron convocatorias nuevas respecto del historial guardado.\n")

    if all_items:
        lines.extend(["## Convocatorias de esta corrida", ""])
        for item in all_items:
            lines.extend(render_item(item))
    return "\n".join(lines).strip() + "\n"


# ─── Exportar Excel ──────────────────────────────────────────────────────────

def export_excel(items: list[Item], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[WARN] openpyxl no está instalado. Saltando exportación Excel.", file=sys.stderr)
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Convocatorias"

    headers = [
        "Revista", "País", "Ciudad", "Región", "Temas", "Tema dossier",
        "Cuartil", "Calidad", "Fuente calidad", "SJR", "H-index",
        "Acceso abierto", "Latindex", "Estado", "Modo", "Urgencia",
        "Cierre (ISO)", "Días restantes", "Puntaje estratégico", "URL",
    ]

    # Encabezados con estilo oscuro
    header_fill = PatternFill("solid", fgColor="0B1020")
    header_font = Font(bold=True, color="00E5FF")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Filas
    alt_fill = PatternFill("solid", fgColor="111827")
    for row_idx, item in enumerate(items, start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        values = [
            item.source_name,
            item.country,
            item.city,
            item.region,
            ", ".join(item.themes),
            item.dossier_topic,
            item.quartile or "—",
            item.quality_label or "—",
            item.quality_source or "—",
            item.sjr or "—",
            item.h_index or "—",
            "Sí" if item.open_access else "No",
            "Sí" if item.latindex_catalogada else "No",
            item.status,
            item.call_mode,
            item.urgency,
            item.deadline_iso or "—",
            item.days_left if item.days_left is not None else "—",
            item.strategic_score,
            item.url,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill
            if col_idx == len(values):  # URL
                cell.hyperlink = item.url
                cell.font = Font(color="00E5FF", underline="single")

    # Anchos de columna aproximados
    col_widths = [32, 14, 16, 16, 28, 38, 9, 12, 20, 8, 9, 10, 10,
                  14, 12, 12, 16, 12, 14, 50]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"[OK] Excel exportado: {path}")


# ─── GitHub Issues ───────────────────────────────────────────────────────────

def create_issue_if_needed(new_items: list[Item], markdown_report: str) -> None:
    if not CREATE_GITHUB_ISSUE or not new_items:
        return
    repo  = os.getenv("GITHUB_REPOSITORY", "").strip()
    token = os.getenv("GITHUB_TOKEN", "").strip()
    if not repo or not token:
        print("[WARN] Faltan GITHUB_REPOSITORY o GITHUB_TOKEN.", file=sys.stderr)
        return
    owner, name = repo.split("/", 1)
    title   = f"Dossier diario | {datetime.now(timezone.utc).date().isoformat()} | {len(new_items)} novedad(es)"
    payload = {"title": title, "body": markdown_report[:65000], "labels": ["dossier", "convocatorias"]}
    resp = requests.post(
        f"https://api.github.com/repos/{owner}/{name}/issues",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept": "application/vnd.github+json",
            "X-GitHub-Api-Version": "2022-11-28",
        },
        json=payload,
        timeout=DEFAULT_TIMEOUT,
    )
    if resp.status_code >= 300:
        print(f"[WARN] No se pudo crear el issue: {resp.status_code}", file=sys.stderr)
    else:
        print("[OK] Issue creado en GitHub.")


# ─── Email ────────────────────────────────────────────────────────────────────

def send_email_if_configured(new_items: list[Item], markdown_report: str) -> None:
    if not new_items:
        return
    smtp_host     = os.getenv("SMTP_HOST", "").strip()
    smtp_port_raw = os.getenv("SMTP_PORT", "465").strip()
    smtp_user     = os.getenv("SMTP_USER", "").strip()
    smtp_password = os.getenv("SMTP_PASSWORD", "").strip()
    email_to      = os.getenv("EMAIL_TO", "").strip()
    if not all([smtp_host, smtp_user, smtp_password, email_to]):
        print("[INFO] Email omitido: faltan variables SMTP.")
        return
    try:
        smtp_port = int(smtp_port_raw)
    except ValueError:
        smtp_port = 465

    # Cuerpo HTML simple
    items_html = "".join(
        f"<tr><td>{i.source_name}</td><td>{i.dossier_topic or i.title}</td>"
        f"<td>{i.quartile or i.quality_label or '—'}</td>"
        f"<td>{i.deadline_iso or 'sin fecha'}</td>"
        f"<td><a href='{i.url}'>Ir</a></td></tr>"
        for i in new_items[:20]
    )
    html_body = f"""
    <html><body style="font-family:sans-serif;background:#0b1020;color:#e6f1ff;padding:20px">
    <h2 style="color:#00e5ff">Dossier diario · {len(new_items)} novedad(es)</h2>
    <table border="1" cellpadding="6" style="border-collapse:collapse;width:100%">
      <thead style="background:#111827">
        <tr><th>Revista</th><th>Tema</th><th>Calidad</th><th>Cierre</th><th>Enlace</th></tr>
      </thead>
      <tbody>{items_html}</tbody>
    </table>
    <p style="color:#8ea2c5;margin-top:20px">Generado por Dossier Tracker Académico v4</p>
    </body></html>
    """

    msg = EmailMessage()
    msg["Subject"] = f"Dossier diario | {len(new_items)} novedad(es) · {datetime.now(timezone.utc).date().isoformat()}"
    msg["From"]    = smtp_user
    msg["To"]      = email_to
    msg.set_content(markdown_report)
    msg.add_alternative(html_body, subtype="html")

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(smtp_host, smtp_port, context=context) as server:
        server.login(smtp_user, smtp_password)
        server.send_message(msg)
    print("[OK] Email enviado.")


# ─── Dashboard JSON ──────────────────────────────────────────────────────────

def item_to_payload(item: Item, is_new: bool) -> dict[str, Any]:
    p = asdict(item)
    p["is_new"] = is_new
    return p


def build_source_catalog(sources: list[dict[str, Any]], all_items: list[Item]) -> list[dict[str, Any]]:
    by_source: dict[str, dict[str, Any]] = {}
    items_by_source: dict[str, list[Item]] = {}
    for item in all_items:
        items_by_source.setdefault(item.source_name, []).append(item)

    for source in sources:
        name         = source.get("name", "")
        source_items = items_by_source.get(name, [])
        has_open      = any(x.call_mode == "deadline" and x.status == "abierta" for x in source_items)
        has_permanent = any(x.call_mode == "permanent" for x in source_items)
        reopen_date   = source.get("reopen_date")

        if source.get("call_mode_hint") == "catalog_only":
            status = "sin convocatoria visible"
        elif has_open:
            status = "activa hoy"
        elif has_permanent or str(source.get("call_mode_hint","")).lower() == "permanent":
            status = "recepcion continua"
        elif reopen_date:
            status = "reapertura anunciada"
        else:
            status = "sin convocatoria visible"

        best_item = max(source_items, key=lambda x: x.strategic_score, default=None)
        by_source[name] = {
            "source_name":    name,
            "url":            source.get("url", ""),
            "region":         source.get("region", ""),
            "country":        source.get("country", ""),
            "city":           source.get("city", ""),
            "language":       source.get("language", ""),
            "themes":         source.get("tags", []),
            "quality_label":  best_item.quality_label if best_item else source.get("quality_label", "—"),
            "quartile":       best_item.quartile       if best_item else source.get("quartile"),
            "quality_source": best_item.quality_source if best_item else source.get("quality_source"),
            "sjr":            best_item.sjr             if best_item else None,
            "h_index":        best_item.h_index         if best_item else None,
            "open_access":    best_item.open_access     if best_item else False,
            "latindex_catalogada": best_item.latindex_catalogada if best_item else False,
            "status":         status,
            "reopen_date":    reopen_date,
            "items_detected": len(source_items),
            "call_type_detected": (
                max(source_items, key=lambda x: x.strategic_score).call_type
                if source_items else "—"
            ),
            "url_confidence": (
                max(source_items, key=lambda x: x.strategic_score).url_confidence
                if source_items else ""
            ),
        }
    return sorted(by_source.values(), key=lambda x: (x["country"], x["source_name"]))


def build_dashboard_payload(
    all_items: list[Item],
    new_items: list[Item],
    sources: list[dict[str, Any]],
    url_cache: dict | None = None,
) -> dict[str, Any]:
    url_cache = url_cache or {}
    new_fps      = {item.fingerprint for item in new_items}
    open_deadline = [x for x in all_items if x.call_mode == "deadline" and x.status == "abierta"]
    permanent     = [x for x in all_items if x.call_mode == "permanent"]
    upcoming      = sorted(open_deadline, key=lambda x: (x.deadline_iso or "9999", x.days_left or 999999))[:10]
    next_deadline = upcoming[0] if upcoming else None

    top_strategic = sorted(
        all_items,
        key=lambda x: (-x.strategic_score, {"Q1":0,"Q2":1,"Q3":2,"Q4":3}.get(x.quartile or "", 9)),
    )[:12]

    theme_counts, country_counts, quartile_counts = {}, {}, {}
    for item in all_items:
        country_counts[item.country] = country_counts.get(item.country, 0) + 1
        for t in item.themes:
            theme_counts[t] = theme_counts.get(t, 0) + 1
        q = item.quartile or item.quality_label or "Sin dato"
        quartile_counts[q] = quartile_counts.get(q, 0) + 1

    source_catalog = build_source_catalog(sources, all_items)

    status_counts = {"activa hoy": 0, "recepcion continua": 0,
                     "reapertura anunciada": 0, "sin convocatoria visible": 0}
    for row in source_catalog:
        status_counts[row["status"]] = status_counts.get(row["status"], 0) + 1

    # Conteos por tipo de convocatoria
    dossier_items  = [x for x in all_items if x.call_type == "dossier"]
    especial_items = [x for x in all_items if x.call_type == "especial"]
    cerrada_items  = [x for x in all_items if x.call_type == "cerrada" or x.status == "vencida"]
    call_type_counts = {}
    for item in all_items:
        ct = item.call_type
        call_type_counts[ct] = call_type_counts.get(ct, 0) + 1

    # Conteos de catálogo por estado
    cat_sin_conv = status_counts.get("sin convocatoria visible", 0)

    metrics = {
        "generated_at":          datetime.now(timezone.utc).isoformat(),
        "total_items":           len(all_items),
        "new_items":             len(new_items),
        "open_with_deadline":    len(open_deadline),
        "permanent_items":       len(permanent),
        "source_count":          len(source_catalog),
        "source_count_scraped":  sum(1 for s in sources if s.get("call_mode_hint") != "catalog_only"),
        "url_confidence_counts": {
            "high":   sum(1 for v in url_cache.values() if isinstance(v,dict) and v.get("confidence")=="high"),
            "medium": sum(1 for v in url_cache.values() if isinstance(v,dict) and v.get("confidence")=="medium"),
            "low":    sum(1 for v in url_cache.values() if isinstance(v,dict) and v.get("confidence")=="low"),
        },
        "active_sources":        status_counts.get("activa hoy", 0),
        "permanent_sources":     status_counts.get("recepcion continua", 0),
        "reopening_sources":     status_counts.get("reapertura anunciada", 0),
        "sin_convocatoria":      cat_sin_conv,
        "dossier_count":         len(dossier_items),
        "especial_count":        len(especial_items),
        "cerrada_count":         len(cerrada_items),
        "call_type_counts":      call_type_counts,
        "next_deadline_title":   next_deadline.title        if next_deadline else None,
        "next_deadline_iso":     next_deadline.deadline_iso if next_deadline else None,
        "next_deadline_url":     next_deadline.url          if next_deadline else None,
    }

    return {
        "metrics":        metrics,
        "items":          [item_to_payload(x, x.fingerprint in new_fps) for x in all_items],
        "upcoming":       [item_to_payload(x, x.fingerprint in new_fps) for x in upcoming],
        "top_strategic":  [item_to_payload(x, x.fingerprint in new_fps) for x in top_strategic],
        "theme_counts":   [{"theme": k, "count": v} for k, v in sorted(theme_counts.items(), key=lambda kv: -kv[1])],
        "country_counts": [{"country": k, "count": v} for k, v in sorted(country_counts.items(), key=lambda kv: -kv[1])],
        "quartile_counts":   [{"label": k, "count": v} for k, v in sorted(quartile_counts.items(), key=lambda kv: -kv[1])],
        "call_type_counts":  [{"type": k, "count": v} for k, v in sorted(call_type_counts.items(), key=lambda kv: -kv[1])],
        "source_catalog":    source_catalog,
    }


# ─── Main ─────────────────────────────────────────────────────────────────────

def main() -> int:
    for d in [REPORTS_DIR, DATA_DIR, DOCS_DATA_DIR]:
        d.mkdir(parents=True, exist_ok=True)

    keywords  = read_keywords()
    sources   = load_yaml(CONFIG_DIR / "sources.yml").get("sources", [])
    if not sources:
        raise RuntimeError("No hay fuentes en config/sources.yml")

    # Cargar bases de calidad
    scimago_db  = load_scimago_db(RANKINGS_DIR)
    latindex_db = load_latindex_db(RANKINGS_DIR)

    session   = get_session()
    all_items: list[Item] = []

    # ── Resolver URLs reales para fuentes SCImago/Latindex ───────────────────
    url_cache = load_url_cache(URL_CACHE_PATH)
    url_cache = resolve_urls_batch(sources, url_cache, session, max_new=400)
    save_url_cache(URL_CACHE_PATH, url_cache)
    resolved_count = sum(1 for v in url_cache.values() if v)
    print(f"[INFO] URLs resueltas en caché: {resolved_count}/{len(url_cache)}")

    # Enriquecer calidad (instantáneo, lee CSV locales)
    enriched = []
    for source in sources:
        qf = enrich_source_quality(source, scimago_db, latindex_db)
        enriched.append((source, qf))

    # Scraping concurrente
    def _scrape_source(args: tuple) -> list[Item]:
        source, quality_fields = args
        try:
            sess = get_session()
            # Anotar confidence en source desde caché
            issn_k = (source.get("issn","") or "").replace("-","").strip()
            if issn_k and url_cache and issn_k in url_cache:
                entry = url_cache[issn_k]
                conf  = entry.get("confidence","") if isinstance(entry,dict) else "medium"
                source = {**source, "_url_confidence": conf}
            items = collect_items(sess, source, keywords, quality_fields, url_cache)
            q_label = quality_fields.get("quartile") or quality_fields.get("quality_label") or "—"
            print(f"[INFO] {source['name']}: {len(items)} ítem(s) · {q_label}")
            return items
        except Exception as exc:
            print(f"[WARN] {source['name']}: {exc}", file=sys.stderr)
            return []

    print(f"[INFO] Iniciando scraping de {len(enriched)} fuentes con {MAX_WORKERS} workers...")
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        results = list(executor.map(_scrape_source, enriched))
    for items in results:
        all_items.extend(items)

    # Deduplicación global
    deduped: dict[str, Item] = {}
    for item in all_items:
        key = canonical_item_key(item)
        current = deduped.get(key)
        if current is None or item_rank(item) > item_rank(current):
            deduped[key] = item
    all_items = sort_items(list(deduped.values()))

    # Historial
    seen_path = DATA_DIR / "seen_items.json"
    seen      = load_seen(seen_path)
    new_items = [item for item in all_items if item.fingerprint not in seen]

    # Reportes
    today = datetime.now(timezone.utc).date().isoformat()
    markdown_report   = render_markdown(new_items, all_items)
    dashboard_payload = build_dashboard_payload(all_items, new_items, sources, url_cache)

    (REPORTS_DIR / f"dossier_{today}.md").write_text(markdown_report, encoding="utf-8")
    (REPORTS_DIR / "latest.md").write_text(markdown_report, encoding="utf-8")

    save_json(DATA_DIR      / "latest_items.json", [asdict(i) for i in all_items])
    save_json(DATA_DIR      / "new_items.json",    [asdict(i) for i in new_items])
    save_json(DOCS_DATA_DIR / "dashboard.json",    dashboard_payload)
    save_json(DOCS_DATA_DIR / "latest_items.json", [asdict(i) for i in all_items])
    save_json(DOCS_DATA_DIR / "new_items.json",    [asdict(i) for i in new_items])
    save_json(seen_path,                           sorted(seen | {i.fingerprint for i in all_items}))

    # Excel
    export_excel(all_items, REPORTS_DIR / f"dossier_{today}.xlsx")
    export_excel(all_items, DOCS_DATA_DIR / "latest.xlsx")

    create_issue_if_needed(new_items, markdown_report)
    send_email_if_configured(new_items, markdown_report)

    print(f"[OK] Total: {len(all_items)} | Nuevas: {len(new_items)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
